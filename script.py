import pandas as pd
import requests
import re
from urllib.parse import urlparse
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Função para extrair o slug do URL
def slug_from_url(url):
    path = urlparse(url).path
    parts = path.strip('/').split('/')
    return parts[-1] if parts[-1] else parts[-2]

# Função para obter preço e estoque via API do WooCommerce
def get_product_info(slug):
    api_url = f'https://meudropbrasil.com/wp-json/wc/store/products?slug={slug}'
    try:
        response = requests.get(api_url, timeout=20)
        products = response.json()

        # Verifica se a resposta é uma lista com ao menos um item
        if isinstance(products, list) and len(products) > 0:
            product = products[0]  # Acessa o primeiro item

            # Verifica se os dados esperados existem
            if 'prices' in product and 'stock_availability' in product:
                # Preço em centavos → converter para reais
                price_cents = product['prices'].get('price', 0)
                price = int(price_cents) / 100 if price_cents else None

                # Estoque: extrai número do texto
                stock_text = product['stock_availability'].get('text', '')
                stock_qty = None
                if stock_text:
                    m = re.search(r'(\d+)', stock_text)
                    stock_qty = int(m.group(1)) if m else None

                return price, stock_qty
            else:
                print(f"Produto encontrado mas com dados incompletos: {slug}")
        else:
            print(f"Produto não encontrado: {slug}")
    except requests.exceptions.RequestException as e:
        print(f'Erro de conexão ao acessar {api_url}: {e}')
    except Exception as e:
        print(f'Erro ao processar {api_url}: {e}')

    return None, None

# Caminho da planilha
caminho_planilha = r'C:\Users\Administrator\Desktop\Vendas Drop Brasil.xlsx'

# Carregar o arquivo com openpyxl
book = load_workbook(caminho_planilha)

# Carregar a aba "Produtos" com pandas
df = pd.read_excel(caminho_planilha, sheet_name='Produtos')

# Processar cada linha da planilha
for index, row in df.iterrows():
    url_produto = row['URL']
    slug = slug_from_url(url_produto)
    preco, estoque = get_product_info(slug)

    # Atualizar apenas se os dados existirem
    if preco is not None:
        df.loc[index, 'Preço de Custo'] = preco
    if estoque is not None:
        df.loc[index, 'Estoque'] = estoque

# Criar novo workbook para salvar as alterações
wb = Workbook()
wb.remove(wb.active)  # Remove aba padrão

# Copiar todas as abas, atualizando a "Produtos"
for sheet_name in book.sheetnames:
    if sheet_name == 'Produtos':
        ws = wb.create_sheet(title='Produtos')
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    else:
        ws = wb.create_sheet(title=sheet_name)
        source_sheet = book[sheet_name]
        for row in source_sheet.iter_rows(values_only=True):
            ws.append(row)

# Salvar o novo arquivo no mesmo caminho
wb.save(caminho_planilha)

print('Planilha atualizada com sucesso!')
